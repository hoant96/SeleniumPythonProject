import openpyxl


class HomePageData:
    test_homepage_data = [{"name" : "demo1", "email" : "test1@gmail.com", "gender" : "Male"}, {"name" : "demo2", "email" : "test2@gmail.com", "gender" : "Female"}]

    @staticmethod
    def getTestData(testcaseName):
        book = openpyxl.load_workbook("/Users/htnguyen/Documents/exampleData.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcaseName:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)

        return [Dict]
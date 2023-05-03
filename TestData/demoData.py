import openpyxl

book = openpyxl.load_workbook("/Users/htnguyen/Documents/exampleData.xlsx")

sheet = book.active
Dict = {}
Dicts = []
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "test1"

sheet.max_row

print(sheet.max_column)
print(sheet['B2'].value)

for i in range(2, sheet.max_row+1):
    for j in range(2, sheet.max_column + 1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    Dicts.append(Dict)

print(Dicts)

